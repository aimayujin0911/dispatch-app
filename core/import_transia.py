#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
トランシア実データインポートスクリプト

車両一覧、従業員名簿、取引先リスト、配車表からデータを取り込む。
Usage: python core/import_transia.py
"""

import sys
import os
import re
import shutil

# core ディレクトリを sys.path に追加
CORE_DIR = os.path.dirname(os.path.abspath(__file__))
if CORE_DIR not in sys.path:
    sys.path.insert(0, CORE_DIR)

# テナントIDを環境変数にセット（database.py が参照する）
os.environ.setdefault("TENANT_ID", "transia")

from database import SessionLocal, engine, Base
from models import Vehicle, Driver, Client, User

TENANT_ID = "transia"

# ========== ソースファイルパス ==========
# OneDrive ファイルはロックされることがあるので、tmp にコピーして使う
TMP_DIR = os.path.join(CORE_DIR, "tmp_import")

VEHICLE_SRC = os.path.join(os.path.expanduser("~"), "OneDrive", "ドキュメント", "車両一覧.xlsx")
EMPLOYEE_SRC = os.path.join(os.path.expanduser("~"), "OneDrive", "ドキュメント", "jinjer_労働者名簿_15459_20260323 (1).csv")
CLIENT_SRC = os.path.join(os.path.expanduser("~"), "OneDrive", "ドキュメント", "取引先_20260324165912.xlsx")
DISPATCH_SRC = os.path.join(os.path.expanduser("~"), "Downloads", "12月　2025車両配車表.xlsx")

VEHICLE_FILE = os.path.join(TMP_DIR, "車両一覧.xlsx")
EMPLOYEE_FILE = os.path.join(TMP_DIR, "従業員名簿.csv")
CLIENT_FILE = os.path.join(TMP_DIR, "取引先.xlsx")
DISPATCH_FILE = os.path.join(TMP_DIR, "配車表.xlsx")


def copy_source_files():
    """OneDrive 等のソースファイルを tmp にコピー（ロック回避）"""
    os.makedirs(TMP_DIR, exist_ok=True)
    sources = [
        (VEHICLE_SRC, VEHICLE_FILE),
        (EMPLOYEE_SRC, EMPLOYEE_FILE),
        (CLIENT_SRC, CLIENT_FILE),
        (DISPATCH_SRC, DISPATCH_FILE),
    ]
    for src, dst in sources:
        if os.path.exists(src):
            shutil.copy2(src, dst)
            print(f"  コピー: {os.path.basename(src)}")
        else:
            print(f"  [警告] ファイルが見つかりません: {src}")


# ========== 車両タイプ解析 ==========

def normalize_text(s):
    """半角カナ→全角カナ、全角英数→半角英数に統一"""
    if not s:
        return ""
    # 半角カナ→全角カナ変換テーブル
    hankaku = "ｱｲｳｴｵｶｷｸｹｺｻｼｽｾｿﾀﾁﾂﾃﾄﾅﾆﾇﾈﾉﾊﾋﾌﾍﾎﾏﾐﾑﾒﾓﾔﾕﾖﾗﾘﾙﾚﾛﾜﾝ"
    zenkaku = "アイウエオカキクケコサシスセソタチツテトナニヌネノハヒフヘホマミムメモヤユヨラリルレロワン"
    hk_map = str.maketrans(hankaku, zenkaku)
    # 濁点・半濁点
    dakuten_map = {
        "ｶﾞ": "ガ", "ｷﾞ": "ギ", "ｸﾞ": "グ", "ｹﾞ": "ゲ", "ｺﾞ": "ゴ",
        "ｻﾞ": "ザ", "ｼﾞ": "ジ", "ｽﾞ": "ズ", "ｾﾞ": "ゼ", "ｿﾞ": "ゾ",
        "ﾀﾞ": "ダ", "ﾁﾞ": "ヂ", "ﾂﾞ": "ヅ", "ﾃﾞ": "デ", "ﾄﾞ": "ド",
        "ﾊﾞ": "バ", "ﾋﾞ": "ビ", "ﾌﾞ": "ブ", "ﾍﾞ": "ベ", "ﾎﾞ": "ボ",
        "ﾊﾟ": "パ", "ﾋﾟ": "ピ", "ﾌﾟ": "プ", "ﾍﾟ": "ペ", "ﾎﾟ": "ポ",
        "ｳﾞ": "ヴ",
    }
    result = s
    for hk, zk in dakuten_map.items():
        result = result.replace(hk, zk)
    # 残りの半角カナ
    result = result.translate(hk_map)
    # 小文字半角カナ
    small_hk = {"ｧ": "ァ", "ｨ": "ィ", "ｩ": "ゥ", "ｪ": "ェ", "ｫ": "ォ",
                "ｬ": "ャ", "ｭ": "ュ", "ｮ": "ョ", "ｯ": "ッ", "ｰ": "ー"}
    for hk, zk in small_hk.items():
        result = result.replace(hk, zk)
    # 全角英数→半角
    zen_alpha = "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"
    han_alpha = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    zen_num = "０１２３４５６７８９"
    han_num = "0123456789"
    zen_lower = "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ"
    han_lower = "abcdefghijklmnopqrstuvwxyz"
    result = result.translate(str.maketrans(zen_alpha + zen_num + zen_lower,
                                            han_alpha + han_num + han_lower))
    return result


def parse_vehicle_type(raw_type):
    """
    車両タイプ文字列を解析して (type, capacity, temperature_zone, has_power_gate) を返す

    例:
      13tW → (ウイング車, 13.0, 常温, False)
      13tWG → (ウイング車, 13.0, 常温, True)
      7tユ → (ユニック車, 7.0, 常温, False)
      4tワイドW → (ウイング車, 4.0, 常温, False)
      4tワイドチルドG → (チルド車, 4.0, チルド, True)
      3tワイドU → (ユニック車, 3.0, 常温, False)
      3tワイドチルド → (チルド車, 3.0, チルド, False)
      2t箱 → (バン, 2.0, 常温, False)
      2t平 → (平ボディ, 2.0, 常温, False)
      10t箱ドラム → (バン, 10.0, 常温, False)
      ハイエース → (バン, 1.0, 常温, False)
      キャラバン → (バン, 1.0, 常温, False)
    """
    norm = normalize_text(raw_type.strip())

    # 特殊車両
    if "ハイエース" in norm or "キャラバン" in norm:
        return ("バン", 1.0, "常温", False)

    # 積載量を抽出
    capacity = 0.0
    cap_match = re.search(r'(\d+)t', norm, re.IGNORECASE)
    if cap_match:
        capacity = float(cap_match.group(1))

    # パワーゲート判定（G が末尾近く、かつ格納/跳上/リンボー含む場合も）
    has_pg = False
    # 「WG」「チルドG」「標準G」「ワイドG」等の G を検出
    # ただし W の直後の G はパワーゲート
    if re.search(r'G(リンボー|格|跳|$|\()', norm):
        has_pg = True
    if re.search(r'WG', norm):
        has_pg = True

    # 温度帯判定
    temp_zone = "常温"
    if "チルド" in norm:
        temp_zone = "チルド"
    elif "冷凍" in norm:
        temp_zone = "冷凍"
    elif "冷蔵" in norm:
        temp_zone = "冷蔵"

    # 車両タイプ判定
    v_type = "その他"
    if "チルド" in norm:
        v_type = "バン"  # チルド車は箱型
    elif "W" in norm.upper() and "ウイング" not in norm:
        # W = ウイング
        v_type = "ウイング車"
    elif "ウイング" in norm:
        v_type = "ウイング車"
    elif "ユ" in norm or "U" in norm.upper():
        # ユ/U = ユニック
        if "箱" not in norm:
            v_type = "ユニック車"
    elif "平" in norm:
        v_type = "平ボディ"
    elif "箱" in norm:
        v_type = "バン"

    # チルド車の場合はタイプ名を「バン」のまま（温度帯で区別）

    return (v_type, capacity, temp_zone, has_pg)


def import_vehicles(db):
    """車両一覧 Excel からインポート"""
    import openpyxl

    if not os.path.exists(VEHICLE_FILE):
        print("[スキップ] 車両一覧ファイルが見つかりません")
        return 0

    wb = openpyxl.load_workbook(VEHICLE_FILE, data_only=True)
    ws = wb.active

    count = 0
    seen_plates = set()

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=4, values_only=True):
        raw_type, plate = row
        if not raw_type or not plate:
            continue

        plate = str(plate).strip()

        # 「26期増車予定」等の未登録車両はスキップ
        if "増車" in plate or "予定" in plate or "納車" in plate:
            continue

        # 重複スキップ
        if plate in seen_plates:
            continue
        seen_plates.add(plate)

        raw_type = str(raw_type).strip()
        v_type, capacity, temp_zone, has_pg = parse_vehicle_type(raw_type)

        # ナンバープレートから短縮番号を抽出（末尾の数字部分）
        num_match = re.search(r'(\d+)\s*$', plate)
        number = plate  # フルプレートを number にする

        vehicle = Vehicle(
            tenant_id=TENANT_ID,
            number=plate,
            type=v_type,
            capacity=capacity,
            temperature_zone=temp_zone,
            has_power_gate=has_pg,
            status="空車",
            notes=f"元データ: {raw_type}",
        )
        db.add(vehicle)
        count += 1

    db.flush()
    print(f"  車両: {count} 台インポート")
    return count


def import_drivers(db):
    """従業員名簿 CSV からインポート"""
    import csv

    if not os.path.exists(EMPLOYEE_FILE):
        print("[スキップ] 従業員名簿ファイルが見つかりません")
        return 0

    # エンコーディングを自動検出（cp932 or utf-8）
    for enc in ["cp932", "utf-8-sig", "utf-8"]:
        try:
            with open(EMPLOYEE_FILE, "r", encoding=enc) as f:
                f.read(100)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    count = 0
    with open(EMPLOYEE_FILE, "r", encoding=enc) as f:
        reader = csv.reader(f)
        header = next(reader)  # ヘッダー行スキップ

        for row in reader:
            if len(row) < 3:
                continue
            emp_no, last_name, first_name = row[0], row[1], row[2]
            if not last_name:
                continue

            full_name = f"{last_name} {first_name}".strip()

            driver = Driver(
                tenant_id=TENANT_ID,
                name=full_name,
                status="待機中",
                notes=f"社員番号: {emp_no}",
            )
            db.add(driver)
            count += 1

    db.flush()
    print(f"  ドライバー: {count} 名インポート")
    return count


def import_clients(db):
    """取引先 Excel からインポート"""
    import openpyxl

    if not os.path.exists(CLIENT_FILE):
        print("[スキップ] 取引先ファイルが見つかりません")
        return 0

    wb = openpyxl.load_workbook(CLIENT_FILE, data_only=True, read_only=True)
    ws = wb.active

    count = 0
    rows = ws.iter_rows(min_row=2, values_only=True)  # ヘッダースキップ

    for row in rows:
        # A: 法人格, B: 設定位置, C: 会社, D: ふりがな, E: 支店
        # G: 郵便番号, H: 都道府県, I: 市区, J: 町村, K: 番地, L: 建物名
        # M: 電話番号, N: FAX, O: 法人番号
        if len(row) < 15:
            continue

        corp_type = str(row[0] or "").strip()  # 株式会社 etc
        position = str(row[1] or "").strip()   # 前/後ろに入れる
        company = str(row[2] or "").strip()
        branch = str(row[4] or "").strip()

        if not company:
            continue

        # 法人格付き名前を組み立て
        if corp_type:
            if "前" in position:
                full_name = f"{corp_type}{company}"
            else:
                full_name = f"{company}{corp_type}"
        else:
            full_name = company

        if branch:
            full_name = f"{full_name} {branch}"

        # 住所組み立て
        postal = str(row[6] or "").strip()
        pref = str(row[7] or "").strip()
        city = str(row[8] or "").strip()
        town = str(row[9] or "").strip()
        street = str(row[10] or "").strip()
        building = str(row[11] or "").strip()

        address_parts = [pref, city, town, street, building]
        address = "".join(p for p in address_parts if p)
        if postal:
            address = f"〒{postal} {address}"

        phone = str(row[12] or "").strip()
        fax = str(row[13] or "").strip()
        tax_id = str(row[14] or "").strip() if row[14] else ""

        client = Client(
            tenant_id=TENANT_ID,
            name=full_name,
            address=address,
            phone=phone,
            fax=fax,
            tax_id=tax_id if tax_id else "",
        )
        db.add(client)
        count += 1

        # メモリ節約: 500件ごとにflush
        if count % 500 == 0:
            db.flush()
            print(f"    ... {count} 件処理中")

    db.flush()
    wb.close()
    print(f"  取引先: {count} 件インポート")
    return count


def match_vehicle_drivers(db):
    """
    配車表から車両番号とドライバー名の紐付けを試みる。
    各課のシートごとにレイアウトが異なるので個別に処理する。
    """
    import openpyxl

    if not os.path.exists(DISPATCH_FILE):
        print("[スキップ] 配車表ファイルが見つかりません")
        return 0

    wb = openpyxl.load_workbook(DISPATCH_FILE, data_only=True)

    # 全車両を取得（ナンバー末尾4桁でマッチング用）
    vehicles = db.query(Vehicle).filter(Vehicle.tenant_id == TENANT_ID).all()
    # 末尾番号 → Vehicle のマップ
    plate_map = {}
    for v in vehicles:
        # ナンバーから末尾数字を抽出
        nums = re.findall(r'\d+', v.number)
        if nums:
            last_num = nums[-1]  # 末尾の数字部分
            plate_map[last_num] = v

    # 全ドライバーを取得（姓でマッチング）
    drivers = db.query(Driver).filter(Driver.tenant_id == TENANT_ID).all()
    driver_map = {}
    for d in drivers:
        # 姓（スペース区切りの最初の部分）
        surname = d.name.split()[0] if d.name else ""
        driver_map[surname] = d

    matched = 0

    # ================================================================
    # 各課のデータ形式（配車表Excelの実データに基づく）
    # ================================================================
    # 1課: Row5="164 島田"(社内番号+名前), Row6="2900G(12本)"(ナンバー末尾+車格)
    #       → Row6の先頭数字でplate_mapマッチ + Row5のドライバー名
    # 2課: Row5="2ｔ標準箱"(車種), Row6="6200.0"(ナンバー末尾)
    #       → Row6の数字でplate_mapマッチ, Row4にドライバー名
    # 3課: Row5="7tユニ"(車種), Row6="春日部102 を7000"(フルナンバー)
    #       → Row6末尾数字でplate_mapマッチ, Row4にドライバー名
    # 4課: Row5="中村"(ドライバー名), Row6="4ｔワイドＷG(格納)"(車種)
    #       → ドライバー名から逆引きで車両特定
    # 5課: Row5="佐藤（武）"(ドライバー名), Row6="3tﾜｲﾄﾞﾁﾙﾄﾞ"(車種)
    #       → ドライバー名から逆引きで車両特定
    # ================================================================

    skip_labels = {"未配車", "修理手配中", "架装手配中", "フリー", "未確定",
                   "荷物情報（受けてない）", "空車情報（変更・キャンセル分）",
                   "未配車 ", "予備車", "ﾄﾞﾗﾑ車両", "修理手配中", "架装手配中"}

    def clean_driver_name(name):
        """ドライバー名をクリーンアップ"""
        if not name:
            return None
        name = re.sub(r'[\(（].*?[\)）]', '', name).strip()
        # 数字のみはスキップ
        if re.match(r'^\d+$', name):
            return None
        if name in skip_labels or not name:
            return None
        return name

    def extract_plate_num(text):
        """テキストからナンバー末尾（3-4桁の数字）を抽出"""
        if not text:
            return None
        text = str(text).strip()
        # ".0" を除去（Excel数値）
        text = re.sub(r'\.0$', '', text)
        # パターン1: 数字のみ（"6200", "886"）
        m = re.match(r'^(\d{3,4})$', text)
        if m:
            return m.group(1)
        # パターン2: "2900G(12本)" → 先頭の数字
        m = re.match(r'^(\d{3,4})', text)
        if m:
            return m.group(1)
        # パターン3: "春日部102 を7000" → 末尾の数字
        nums = re.findall(r'\d+', text)
        if nums:
            return nums[-1]
        return None

    for sheet_name in ["1課", "2課", "3課", "4課", "5課"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        print(f"  配車表 [{sheet_name}] 解析中...")

        sheet_matched = 0
        for col in range(3, ws.max_column + 1):
            r4 = str(ws.cell(4, col).value).strip() if ws.cell(4, col).value else ""
            r5 = str(ws.cell(5, col).value).strip() if ws.cell(5, col).value else ""
            r6 = str(ws.cell(6, col).value).strip() if ws.cell(6, col).value else ""

            if not r5:
                continue
            # スキップ対象ラベル
            if r5 in skip_labels:
                continue

            driver_name = None
            vehicle_num = None
            vehicle = None
            driver = None

            if sheet_name == "1課":
                # Row5="164 島田" → ドライバー名, Row6="2900G(12本)" → ナンバー末尾
                parts = r5.split()
                if len(parts) >= 2:
                    driver_name = clean_driver_name(parts[1])
                elif not r5[0].isdigit():
                    driver_name = clean_driver_name(r5)
                vehicle_num = extract_plate_num(r6)

            elif sheet_name == "2課":
                # Row5="2ｔ標準箱"(車種), Row6="6200.0"(ナンバー末尾), Row4=ドライバー名
                driver_name = clean_driver_name(r4) if r4 else None
                vehicle_num = extract_plate_num(r6)

            elif sheet_name == "3課":
                # Row5="7tユニ"(車種), Row6="春日部102 を7000"(フルナンバー), Row4=ドライバー名
                driver_name = clean_driver_name(r4) if r4 else None
                vehicle_num = extract_plate_num(r6)

            elif sheet_name in ("4課", "5課"):
                # Row5="中村"(ドライバー名), Row6="4ｔワイドＷG(格納)"(車種), Row7=ナンバー末尾
                driver_name = clean_driver_name(r5)
                r7 = str(ws.cell(7, col).value).strip() if ws.cell(7, col).value else ""
                vehicle_num = extract_plate_num(r7)

            # --- マッチング ---
            if vehicle_num:
                vehicle = plate_map.get(vehicle_num)
            if driver_name:
                driver = driver_map.get(driver_name)

            # 課の情報をセット
            if vehicle:
                vehicle.department = sheet_name
            if driver:
                driver.department = sheet_name

            if vehicle and driver:
                vehicle.default_driver_id = driver.id
                sheet_matched += 1
                matched += 1
            elif vehicle and not driver:
                # 車両のみマッチ（ドライバー不明）→ 課だけセット済み
                pass
            elif driver and not vehicle:
                # ドライバーのみマッチ（4課・5課等）→ 逆引きで車両にもdepartmentセット
                for v in vehicles:
                    if v.default_driver_id == driver.id:
                        v.department = sheet_name
                        sheet_matched += 1
                        matched += 1
                        break

        print(f"    → {sheet_name}: {sheet_matched} 件マッチ")

    # 最終パス: departmentが設定されたドライバーの担当車両にも課をセット
    for v in vehicles:
        if not v.department and v.default_driver_id:
            for d in drivers:
                if d.id == v.default_driver_id and d.department:
                    v.department = d.department
                    break

    db.flush()
    print(f"  車両-ドライバー紐付け: {matched} 件")
    return matched


def delete_existing_data(db):
    """既存の transia データを削除（admin/dispatcher ユーザーは残す）"""
    # 車両に紐づく配車を先に削除（外部キー制約対応）
    from models import Dispatch, Shipment, Attendance, DailyReport

    print("既存データ削除中...")

    # 配車データ削除
    del_dispatches = db.query(Dispatch).filter(Dispatch.tenant_id == TENANT_ID).delete()
    print(f"  配車: {del_dispatches} 件削除")

    # 車両削除
    del_vehicles = db.query(Vehicle).filter(Vehicle.tenant_id == TENANT_ID).delete()
    print(f"  車両: {del_vehicles} 件削除")

    # ドライバー削除（driver ロールの User も削除）
    driver_users = db.query(User).filter(
        User.tenant_id == TENANT_ID,
        User.role == "driver"
    ).delete()
    print(f"  ドライバーユーザー: {driver_users} 件削除")

    del_drivers = db.query(Driver).filter(Driver.tenant_id == TENANT_ID).delete()
    print(f"  ドライバー: {del_drivers} 件削除")

    # 取引先削除
    del_clients = db.query(Client).filter(Client.tenant_id == TENANT_ID).delete()
    print(f"  取引先: {del_clients} 件削除")

    db.flush()


def main():
    print("=" * 60)
    print("トランシア実データインポート")
    print("=" * 60)

    # テーブル作成（存在しない場合）
    Base.metadata.create_all(bind=engine)

    # ソースファイルをコピー
    print("\nソースファイルをコピー中...")
    copy_source_files()

    db = SessionLocal()
    try:
        # 既存データ削除
        print()
        delete_existing_data(db)
        db.commit()

        # 車両インポート
        print("\n[1/4] 車両インポート...")
        import_vehicles(db)

        # ドライバーインポート
        print("\n[2/4] ドライバーインポート...")
        import_drivers(db)

        # 取引先インポート
        print("\n[3/4] 取引先インポート...")
        import_clients(db)

        # コミット（紐付け前にIDを確定）
        db.commit()

        # 車両-ドライバー紐付け
        print("\n[4/4] 車両-ドライバー紐付け...")
        match_vehicle_drivers(db)

        db.commit()

        # サマリー
        v_count = db.query(Vehicle).filter(Vehicle.tenant_id == TENANT_ID).count()
        d_count = db.query(Driver).filter(Driver.tenant_id == TENANT_ID).count()
        c_count = db.query(Client).filter(Client.tenant_id == TENANT_ID).count()
        assigned = db.query(Vehicle).filter(
            Vehicle.tenant_id == TENANT_ID,
            Vehicle.default_driver_id.isnot(None)
        ).count()

        print("\n" + "=" * 60)
        print("インポート完了!")
        print(f"  車両:             {v_count} 台")
        print(f"  ドライバー:       {d_count} 名")
        print(f"  取引先:           {c_count} 件")
        print(f"  車両-ドライバー:  {assigned} 件紐付済")
        print("=" * 60)

    except Exception as e:
        db.rollback()
        print(f"\n[エラー] インポート失敗: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
